#!/usr/bin/env python3
"""Ingest a Google Search Console "Performance on Search" .xlsx export into seo/data/seo.db.

Usage:
    python ingest_sc.py <path/to/export.xlsx> [--force]

The .xlsx must be exported via Search Console → Performance → Export → Excel.
Expected sheets: Chart, Queries, Pages, Countries, Devices, Filters.
Week window is read from the Filters sheet ("Date" row).
"""

import argparse
import re
import sqlite3
import sys
from datetime import datetime
from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parents[2]
DB_PATH = REPO_ROOT / "seo" / "data" / "seo.db"


SHEET_TABLE_MAP = {
    "Queries":   ("sc_queries",   "query"),
    "Pages":     ("sc_pages",     "page"),
    "Countries": ("sc_countries", "country"),
    "Devices":   ("sc_devices",   "device"),
}


def window_from_chart(ws_chart) -> tuple[str, str]:
    """Derive (week_start, week_end) from the min/max dates in the Chart sheet.

    Fallback for relative-range exports ("Last 7 days") where the Filters sheet
    carries no explicit dates but the per-day Chart rows do.
    """
    dates = [d for d, *_ in load_chart_rows(ws_chart)]
    if not dates:
        raise ValueError("Chart sheet has no dated rows to derive a window from")
    return min(dates), max(dates)


def parse_window(ws_filters, ws_chart=None) -> tuple[str, str]:
    """Read the date range from the Filters sheet, return (week_start, week_end) ISO dates.

    If the Filters "Date" row is a relative phrase ("Last 7 days" etc.) rather than
    an explicit range, fall back to the Chart sheet's min/max dates.
    """
    for row in ws_filters.iter_rows(values_only=True):
        if row and row[0] == "Date":
            raw = row[1]  # e.g. "14 May 2026-20 May 2026" or "14 May 2026 - 20 May 2026"
            m = re.match(r"\s*(.+?)\s*-\s*(.+?)\s*$", raw)
            if m:
                try:
                    start = datetime.strptime(m.group(1), "%d %b %Y").date().isoformat()
                    end   = datetime.strptime(m.group(2), "%d %b %Y").date().isoformat()
                    return start, end
                except ValueError:
                    pass  # not an explicit range — fall through to Chart fallback
            if ws_chart is not None:
                return window_from_chart(ws_chart)
            raise ValueError(f"Cannot parse date range from Filters: {raw!r}")
    if ws_chart is not None:
        return window_from_chart(ws_chart)
    raise ValueError("No 'Date' row in Filters sheet")


def load_sheet_rows(ws):
    """Yield (key, clicks, impressions, ctr, position) for each data row, skipping header."""
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        if row[0] is None:
            continue
        key, clicks, impr, ctr, pos = row
        yield (
            str(key).strip(),
            int(clicks or 0),
            int(impr or 0),
            float(ctr or 0.0),
            float(pos or 0.0),
        )


def load_chart_rows(ws):
    """Yield (date, clicks, impressions, ctr, position) from the Chart sheet."""
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            continue
        date_val, clicks, impr, ctr, pos = row
        if date_val is None:
            continue
        # date_val may be a datetime or a string like "2026-05-14"
        if hasattr(date_val, "date"):
            date_iso = date_val.date().isoformat()
        else:
            date_iso = str(date_val)[:10]
        yield (
            date_iso,
            int(clicks or 0),
            int(impr or 0),
            float(ctr or 0.0),
            float(pos or 0.0),
        )


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("xlsx", help="Path to the Search Console .xlsx export")
    ap.add_argument("--force", action="store_true",
                    help="Replace an existing snapshot for the same window")
    args = ap.parse_args()

    xlsx_path = Path(args.xlsx).expanduser().resolve()
    if not xlsx_path.exists():
        print(f"ERROR: file not found: {xlsx_path}", file=sys.stderr)
        return 2

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if "Filters" not in wb.sheetnames:
        print("ERROR: not a valid SC export (no Filters sheet)", file=sys.stderr)
        return 2

    week_start, week_end = parse_window(wb["Filters"], wb["Chart"] if "Chart" in wb.sheetnames else None)
    print(f"Window: {week_start} -> {week_end}")
    print(f"Source: {xlsx_path}")

    conn = sqlite3.connect(DB_PATH)
    conn.execute("PRAGMA foreign_keys = ON")
    cur = conn.cursor()

    existing = cur.execute(
        "SELECT id FROM snapshots WHERE week_start=? AND week_end=?",
        (week_start, week_end),
    ).fetchone()

    if existing:
        if not args.force:
            print(f"ERROR: snapshot for {week_start}..{week_end} already exists (id={existing[0]}). Use --force to replace.", file=sys.stderr)
            return 3
        cur.execute("DELETE FROM snapshots WHERE id=?", (existing[0],))
        print(f"Removed existing snapshot id={existing[0]}")

    cur.execute(
        "INSERT INTO snapshots (week_start, week_end, source_file) VALUES (?, ?, ?)",
        (week_start, week_end, str(xlsx_path)),
    )
    snapshot_id = cur.lastrowid
    print(f"Created snapshot id={snapshot_id}")

    total = 0
    for sheet_name, (table, key_col) in SHEET_TABLE_MAP.items():
        if sheet_name not in wb.sheetnames:
            print(f"  WARN: sheet {sheet_name} missing, skipping")
            continue
        rows = list(load_sheet_rows(wb[sheet_name]))
        cur.executemany(
            f"INSERT INTO {table} (snapshot_id, {key_col}, clicks, impressions, ctr, position) "
            f"VALUES (?, ?, ?, ?, ?, ?)",
            [(snapshot_id, *r) for r in rows],
        )
        print(f"  {table}: {len(rows)} rows")
        total += len(rows)

    if "Chart" in wb.sheetnames:
        rows = list(load_chart_rows(wb["Chart"]))
        cur.executemany(
            "INSERT INTO sc_daily (snapshot_id, date, clicks, impressions, ctr, position) "
            "VALUES (?, ?, ?, ?, ?, ?)",
            [(snapshot_id, *r) for r in rows],
        )
        print(f"  sc_daily: {len(rows)} rows")
        total += len(rows)

    conn.commit()
    conn.close()
    print(f"Done. {total} rows inserted across all tables.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
